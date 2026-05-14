from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "data" / "ground_truth.xlsx"

wb = Workbook()

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
wrap = Alignment(wrap_text=True, vertical="top")


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
    for r in rows:
        ws.append(r)
    for col_idx, h in enumerate(headers, 1):
        max_len = max([len(str(h))] + [len(str(r[col_idx - 1])) for r in rows])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 12), 60)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = wrap


# Customer
ws = wb.active
ws.title = "Customer"
write_sheet(ws, ["Field", "Value"], [
    ["Account number", "DHL001"],
    ["Company name", "Mind Your Business Inc."],
    ["Primary contact", "Mr Anthony Gressive"],
    ["Alternate contact", "Mrs Anna Gressive"],
    ["Direct phone", "9813640644"],
    ["Registered email", "Anthony@mybiz.com"],
    ["Billing city", "Mumbai"],
    ["Payment terms", "30 days"],
    ["Language preferences", "English, Hinglish"],
])

# Invoices
ws = wb.create_sheet("Invoices")
write_sheet(ws,
    ["Invoice number", "Invoice type", "Amount (INR)", "Currency", "Invoice date", "Due date", "Overdue days", "History"],
    [
        ["DHL123456", "Duty Import", 13600, "INR", "2026-01-01", "2026-01-31", 60,
         "Customer earlier raised a price-list vs invoice mismatch on 1 waybill. Issue resolved; credit note issued. No further open dispute on this invoice — payment is simply pending."],
        ["DHL654321", "Freight Export", 34650, "INR", "2026-02-01", "2026-02-22", 45,
         "Customer logged an issue for 2 delayed shipments. Issue resolved; credit note issued some time ago. Week after the credit note, DHL called the customer; customer confirmed receipt of the credit note. No further open dispute — payment is pending."],
        ["DHL332241", "Freight Import", 9670, "INR", "2026-03-01", "2026-04-07", 30,
         "No prior dispute logged. Standard freight import invoice, simply unpaid past the due date."],
    ])

# Totals
ws = wb.create_sheet("Totals")
write_sheet(ws, ["Field", "Value"], [
    ["Number of invoices", 3],
    ["Total outstanding (INR)", 57920],
])

# Allowed values
ws = wb.create_sheet("AllowedValues")
write_sheet(ws, ["Category", "Value"], [
    ["Amount (INR)", 9670],
    ["Amount (INR)", 13600],
    ["Amount (INR)", 34650],
    ["Amount (INR)", 57920],
    ["Overdue days", 30],
    ["Overdue days", 45],
    ["Overdue days", 60],
    ["Date", "2026-01-01"],
    ["Date", "2026-01-31"],
    ["Date", "2026-02-01"],
    ["Date", "2026-02-22"],
    ["Date", "2026-03-01"],
    ["Date", "2026-04-07"],
])

# Payment methods
ws = wb.create_sheet("PaymentMethods")
write_sheet(ws, ["ID", "Label", "Details"], [
    ["mybill", "DHL MyBill self-serve portal",
     "Customer logs in at mybill.dhl.com with their registered email and password, picks the invoice, and pays online. This is the primary self-serve channel and the same portal where invoice copies can be downloaded."],
    ["virtual_account", "Virtual Account Number (bank transfer)",
     "Each DHL customer has a unique Virtual Account Number tied to their DHL account. Transfer the exact invoice amount to that VAN via NEFT or RTGS — DHL's bank auto-reconciles the payment to this account number. The VAN can be shared by the collections desk if the customer does not have it."],
])

# Policy constants
ws = wb.create_sheet("PolicyConstants")
write_sheet(ws, ["Constant", "Value", "Notes"], [
    ["Promise-to-pay window", "2 business days",
     "Any date the customer offers further out must be politely refused; ask for a tighter date inside the window."],
    ["Soft monthly collection target", "25th of the month", "Try your best to secure payment before this day."],
    ["Proof-of-payment email", "yogesh.jhamb@dhl.com",
     "Where customers email transaction reference + date when claiming \"already paid\"."],
])

# Call dispositions
ws = wb.create_sheet("CallDispositions")
write_sheet(ws, ["Disposition"], [["refusal"], ["reason"], ["promise-to-pay"], ["dispute"], ["escalation"]])

# Escalation contact
ws = wb.create_sheet("EscalationContact")
write_sheet(ws, ["Field", "Value"], [
    ["Name", "Ms Sanorita"],
    ["Designation", "Collections Executive"],
    ["Phone", "09416340644"],
])

wb.save(OUT)
print(f"wrote {OUT}")
